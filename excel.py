# import openpyxl and tkinter modules
import re
import streamlit as st
from openpyxl import *
from tkinter import *
from tkinter import messagebox
# globally declare wb and sheet variable

st.title('projects')
# opening the existing excel file
wb = load_workbook('C:\\Users\\kgovin1x\\OneDrive - Intel Corporation\\Desktop\\excel.xlsx')

# create the sheet object
sheet = wb.active


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    #sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 40
    #sheet.column_dimensions['G'].width = 50

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Oneboxnumber"
    sheet.cell(row=1, column=3).value = "Benchlocation"
    #sheet.cell(row=1, column=4).value = "Form Number"
    sheet.cell(row=1, column=4).value = "Contact Number"
    sheet.cell(row=1, column=5).value = "Email id"
    #sheet.cell(row=1, column=7).value = "Address"


# Function to set focus (cursor)
def focus1(event):
    # set focus on the course_field box
    Oneboxnumber_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the sem_field box
    Benchlocation_field.focus_set()


# Function to set focus
#def focus3(event):
    # set focus on the form_no_field box
    #form_no_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the contact_no_field box
    contact_no_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the email_id_field box
    email_id_field.focus_set()


# Function to set focus
#def focus6(event):
    # set focus on the address_field box
    #address_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    name_field.delete(0, END)
    Oneboxnumber_field.delete(0, END)
    Benchlocation_field.delete(0, END)
    #form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    #address_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
    # if user not fill any entry
    # then print "empty input"
        #for i in range(1, 100):
    if (name_field.get() == "" or
            Oneboxnumber_field.get() == "" or
            Benchlocation_field.get() == "" or
            #Oneboxnumber_field.get() == r'^[a-zA-Z]{3}\d{4}$'
            #form_no_field.get() == "" and
            contact_no_field.get() == "" or
            email_id_field.get() == "" ):
            #address_field.get() == ""):
            #print("empty input")
            messagebox.showerror(title="Message", message="Empty input/pls check the onebox number on the side sticker Example MVX0001")
    elif (re.match(r'^[a-zA-Z]{3}\d{4}$',Oneboxnumber_field.get()) and
          re.match(r'^[a-zA-Z]{3}\d{4}$',contact_no_field.get())):
          messagebox.showerror(title="message", message="Wrong oneboxnumber or contact")
    else:
        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = Oneboxnumber_field.get()
        sheet.cell(row=current_row + 1, column=3).value = Benchlocation_field.get()
        #sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=4).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = email_id_field.get()
        #sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        # save the file
        wb.save('C:\\Users\\kgovin1x\\OneDrive - Intel Corporation\\Desktop\\excel.xlsx')

        # set focus on the name_field box
        name_field.focus_set()

        # call the clear() function
        clear()
        messagebox.showinfo(title="Message",message="Successfully Registered")

# Driver code
if __name__ == "__main__":
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light green')

    # set the title of GUI window
    root.title("registration form")

    # set the configuration of GUI window
    root.geometry("500x300")

    excel()

    # create a Form label
    heading = Label(root, text="ONEBOX & USER DETAILS", bg="light green")

    # create a Name label
    name = Label(root, text="Name", bg="light green")

    # create a Course label
    Oneboxnumber = Label(root, text="Oneboxnumber", bg="light green")

    # create a Semester label
    Benchlocation = Label(root, text="Benchlocation", bg="light green")

    # create a Form No. label
    #form_no = Label(root, text="Form No.", bg="light green")

    # create a Contact No. label
    contact_no = Label(root, text="Contact No.", bg="light green")

    # create a Email id label
    email_id = Label(root, text="Email id", bg="light green")

    # create a address label
    #address = Label(root, text="Address", bg="light green")

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    Oneboxnumber.grid(row=2, column=0)
    Benchlocation.grid(row=3, column=0)
    #form_no.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    #address.grid(row=7, column=0)

    # create a text entry box
    # for typing the information
    name_field = Entry(root)
    Oneboxnumber_field = Entry(root)
    Benchlocation_field = Entry(root)
    #form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    #address_field = Entry(root)

    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    name_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    Oneboxnumber_field.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    Benchlocation_field.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    #form_no_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    contact_no_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus6 function
    #email_id_field.bind("<Return>", focus5)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    name_field.grid(row=1, column=1, ipadx="100")
    Oneboxnumber_field.grid(row=2, column=1, ipadx="100")
    Benchlocation_field.grid(row=3, column=1, ipadx="100")
    #form_no_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    #address_field.grid(row=7, column=1, ipadx="100")


    # call excel function
    excel()
    #def insert():
        #if Oneboxnumber_field.get() ==
     #       messagebox.showerror(title="Message", message="Error in reading the Oneboxnumber")
        #else:
     #       messagebox.showinfo(title="Message",message="Successfully Registered")

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                    bg="Red", command=insert)

    submit.grid(row=8, column=1)

    def app():
        print("installing app")


    Button(root, text="Install the onebox app", fg="Black", bg="grey", command=app).grid(row=9, column=1)
    # start the GUI
    root.mainloop()


    def app():

        print("installing app")
    Button(root, text="Install the onebox app", fg="Black", bg="grey", command=app).grid(row=9, column=1)
    #messagebox.showinfo(title="Message", message="Make sure you have filled all the details")

